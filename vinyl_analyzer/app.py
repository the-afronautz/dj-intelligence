"""Flask backend for the Vinyl Analyzer.

Routes
------
GET  /                  -> single-page UI
POST /api/analyze       -> multipart audio upload, returns BPM/key/Camelot
POST /api/tracks        -> persist a new row (analysis result + user metadata)
GET  /api/tracks        -> list rows (supports ?q= search, ?sort= column, ?order=asc|desc)
DELETE /api/tracks/<id> -> remove a row
GET  /api/export        -> download the full log as a .xlsx file
"""

from __future__ import annotations

import io
import os
import sqlite3
from datetime import datetime
from pathlib import Path

from flask import Flask, g, jsonify, render_template, request, send_file, abort

from analyzer import analyze_bytes

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
DATA_DIR.mkdir(exist_ok=True)
DB_PATH = DATA_DIR / "tracks.db"

ALLOWED_SORT_COLUMNS = {
    "id", "captured_at", "track_name", "artist", "bpm",
    "camelot", "key", "bpm_confidence", "key_confidence",
}

app = Flask(__name__, template_folder="templates", static_folder="static")
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024  # 50 MB per snippet


# ---------------------------------------------------------------------------
# DB helpers
# ---------------------------------------------------------------------------

def get_db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(exc):
    db = g.pop("db", None)
    if db is not None:
        db.close()


def init_db() -> None:
    with sqlite3.connect(DB_PATH) as conn:
        conn.executescript(
            """
            CREATE TABLE IF NOT EXISTS tracks (
                id              INTEGER PRIMARY KEY AUTOINCREMENT,
                captured_at     TEXT NOT NULL,
                track_name      TEXT,
                artist          TEXT,
                bpm             REAL,
                bpm_confidence  REAL,
                key             TEXT,
                camelot         TEXT,
                key_confidence  REAL,
                alt_key         TEXT,
                alt_camelot     TEXT,
                alt_share       REAL,
                alt_bpm         REAL,
                alt_bpm_share   REAL,
                duration_s      REAL,
                notes           TEXT
            );
            CREATE INDEX IF NOT EXISTS idx_tracks_captured_at ON tracks(captured_at DESC);
            CREATE INDEX IF NOT EXISTS idx_tracks_camelot ON tracks(camelot);
            """
        )
        # Migrate old DBs that pre-date the alt_* columns
        cur = conn.execute("PRAGMA table_info(tracks)")
        cols = {row[1] for row in cur.fetchall()}
        for col, decl in [
            ("alt_key", "TEXT"),
            ("alt_camelot", "TEXT"),
            ("alt_share", "REAL"),
            ("alt_bpm", "REAL"),
            ("alt_bpm_share", "REAL"),
        ]:
            if col not in cols:
                conn.execute(f"ALTER TABLE tracks ADD COLUMN {col} {decl}")


# ---------------------------------------------------------------------------
# Routes
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/api/analyze", methods=["POST"])
def analyze():
    """Receive an audio blob (multipart field 'audio') and run analysis.

    Does NOT persist anything — the UI will follow up with /api/tracks
    once the user fills in track name / artist.
    """
    if "audio" not in request.files:
        return jsonify({"error": "no 'audio' file in request"}), 400
    f = request.files["audio"]
    data = f.read()
    if not data:
        return jsonify({"error": "empty audio payload"}), 400

    # Browser MediaRecorder usually gives audio/webm; allow the client
    # to override via a hidden 'suffix' field if needed.
    suffix = request.form.get("suffix") or _suffix_for(f.mimetype, f.filename)

    try:
        result = analyze_bytes(data, suffix=suffix)
    except Exception as exc:  # noqa: BLE001 — surface to client for now
        app.logger.exception("analysis failed")
        return jsonify({"error": f"analysis failed: {exc}"}), 500

    return jsonify(result.to_dict())


@app.route("/api/tracks", methods=["POST"])
def create_track():
    payload = request.get_json(force=True, silent=True) or {}

    captured_at = payload.get("captured_at") or datetime.utcnow().isoformat(timespec="seconds") + "Z"

    db = get_db()
    cur = db.execute(
        """
        INSERT INTO tracks (
            captured_at, track_name, artist, bpm, bpm_confidence,
            key, camelot, key_confidence, alt_key, alt_camelot, alt_share,
            alt_bpm, alt_bpm_share, duration_s, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (
            captured_at,
            (payload.get("track_name") or "").strip() or None,
            (payload.get("artist") or "").strip() or None,
            payload.get("bpm"),
            payload.get("bpm_confidence"),
            payload.get("key"),
            payload.get("camelot"),
            payload.get("key_confidence"),
            payload.get("alt_key"),
            payload.get("alt_camelot"),
            payload.get("alt_share"),
            payload.get("alt_bpm"),
            payload.get("alt_bpm_share"),
            payload.get("duration_seconds"),
            (payload.get("notes") or "").strip() or None,
        ),
    )
    db.commit()
    return jsonify({"id": cur.lastrowid}), 201


@app.route("/api/tracks", methods=["GET"])
def list_tracks():
    q = (request.args.get("q") or "").strip()
    sort = request.args.get("sort", "captured_at")
    order = request.args.get("order", "desc").lower()
    if sort not in ALLOWED_SORT_COLUMNS:
        sort = "captured_at"
    if order not in {"asc", "desc"}:
        order = "desc"

    sql = "SELECT * FROM tracks"
    params: list = []
    if q:
        sql += (
            " WHERE track_name LIKE ? OR artist LIKE ? OR camelot LIKE ?"
            " OR key LIKE ? OR notes LIKE ?"
        )
        like = f"%{q}%"
        params = [like, like, like, like, like]
    sql += f" ORDER BY {sort} {order.upper()}, id DESC LIMIT 1000"

    rows = get_db().execute(sql, params).fetchall()
    return jsonify([dict(r) for r in rows])


@app.route("/api/tracks/<int:track_id>", methods=["DELETE"])
def delete_track(track_id: int):
    db = get_db()
    cur = db.execute("DELETE FROM tracks WHERE id = ?", (track_id,))
    db.commit()
    if cur.rowcount == 0:
        abort(404)
    return ("", 204)


@app.route("/api/export", methods=["GET"])
def export_xlsx():
    """Stream the full log as a freshly built .xlsx."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    rows = get_db().execute(
        "SELECT id, captured_at, track_name, artist, bpm, camelot, key,"
        " bpm_confidence, key_confidence, alt_camelot, alt_key, alt_share,"
        " alt_bpm, alt_bpm_share, duration_s, notes"
        " FROM tracks ORDER BY captured_at DESC, id DESC"
    ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Vinyl Log"

    headers = [
        "ID", "Captured (UTC)", "Track", "Artist", "BPM", "Camelot",
        "Key", "BPM Conf.", "Key Conf.", "Alt BPM", "Alt BPM Score",
        "Alt Camelot", "Alt Key", "Alt Key Share", "Duration (s)", "Notes",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="333333")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["id"], r["captured_at"], r["track_name"], r["artist"],
            r["bpm"], r["camelot"], r["key"],
            r["bpm_confidence"], r["key_confidence"],
            r["alt_bpm"], r["alt_bpm_share"],
            r["alt_camelot"], r["alt_key"], r["alt_share"],
            r["duration_s"], r["notes"],
        ])

    # Column widths
    widths = [6, 22, 30, 24, 8, 10, 14, 10, 10, 10, 12, 12, 14, 10, 12, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = f"vinyl_log_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=fname,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _suffix_for(mimetype: str | None, filename: str | None) -> str:
    if filename and "." in filename:
        return "." + filename.rsplit(".", 1)[-1].lower()
    if not mimetype:
        return ".webm"
    mt = mimetype.lower()
    if "webm" in mt:
        return ".webm"
    if "ogg" in mt:
        return ".ogg"
    if "wav" in mt:
        return ".wav"
    if "mp4" in mt or "m4a" in mt or "aac" in mt:
        return ".m4a"
    if "mpeg" in mt or "mp3" in mt:
        return ".mp3"
    if "flac" in mt:
        return ".flac"
    return ".webm"


# ---------------------------------------------------------------------------
# Entrypoint
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    init_db()
    port = int(os.environ.get("PORT", "5057"))
    # 127.0.0.1 only — no remote access by default.
    app.run(host="127.0.0.1", port=port, debug=False)
